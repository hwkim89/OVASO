import torch
import torch.nn as nn
import numpy as np
import os
import time

from openpyxl import Workbook
from torchvision import datasets, models

def get_model(model_dir, device):
    # Create the covid binary model
    covid_ft = models.resnet50(pretrained=True)
    covid_num_ftrs = covid_ft.fc.in_features
    covid_ft.fc = nn.Linear(covid_num_ftrs, 2)
    covid_ft.load_state_dict(torch.load(f"{model_dir}/covid_binary2.pt")) # covid_binary2는 어디에?
    covid_ft.eval()
    covid_ft = covid_ft.to(device)
    
    # Create the pneumonia binary model
    pneumonia_ft = models.resnet50(pretrained=True)
    pneumonia_num_ftrs = pneumonia_ft.fc.in_features
    pneumonia_ft.fc = nn.Linear(pneumonia_num_ftrs, 2)
    pneumonia_ft.load_state_dict(torch.load(f"{model_dir}/pneumonia_binary2.pt"))
    pneumonia_ft.eval()
    pneumonia_ft = pneumonia_ft.to(device)
    
    # Create the normal binary model
    normal_ft = models.resnet50(pretrained=True)
    normal_num_ftrs = normal_ft.fc.in_features
    normal_ft.fc = nn.Linear(normal_num_ftrs, 2)
    normal_ft.load_state_dict(torch.load(f"{model_dir}/normal_binary2.pt"))
    normal_ft.eval()
    normal_ft = normal_ft.to(device)
    
    # Create the loss function
    criterion = nn.CrossEntropyLoss()
    
    return covid_ft, pneumonia_ft, normal_ft, criterion

def eval_model(dataloaders, covid_model, pneumonia_model, normal_model,
               criterion, batch_size, device, out_dir):
    
    def create_workbook():
        wb = Workbook()      # 워크북을 생성한다.
        ws = wb.active       # 워크 시트를 얻는다.

        ws['A1'] = 'ResNet50'
        ws['B1'] = 'Val ACC'

        ws['D1'] = 'Covid ACC'
        ws['E1'] = 'Covid Recall'
        ws['F1'] = 'Covid Precision'
        ws['G1'] = 'Covid F1'

        ws['I1'] = 'Normal ACC'
        ws['J1'] = 'Normal Recall'
        ws['K1'] = 'Normal Precision'
        ws['L1'] = 'Normal F1'

        ws['N1'] = 'Pneumonia ACC'
        ws['O1'] = 'Pneumonia Recall'
        ws['P1'] = 'Pneumonia Precision'
        ws['Q1'] = 'Pneumonia F1'
        
        return wb, ws
    
    since = time.time()
    
    # Create the workbook and worksheet
    wb, ws = create_workbook()

    # Init parameters
    val_covid_TP, val_normal_TP, val_pneumonia_TP = 0.0, 0.0, 0.0
    val_covid_FN, val_normal_FN, val_pneumonia_FN = 0.0, 0.0, 0.0
    val_covid_TN, val_normal_TN, val_pneumonia_TN = 0.0, 0.0, 0.0
    val_covid_FP, val_normal_FP, val_pneumonia_FP = 0.0, 0.0, 0.0
    
    running_corrects = 0.0
    covid_threshold = 0.95

    print('Evaluation Result')
    print('-' * 10)

    # Each epoch has a training and validation phase
    pneumonia_model.eval()   # Set model to evaluate mode
    covid_model.eval()

    # Iterate over data.
    for inputs, labels in dataloaders:
        inputs = inputs.to(device)
        labels = labels.to(device)

        correct_label = labels.data[0].item()

#         # Previous OVR model
#         pneumonia_outputs = pneumonia_model(inputs)
#         _, pneumonia_preds = torch.max(pneumonia_outputs, 1)
#         binary_label = pneumonia_preds.item()  #pneumonia_label = 0 = normal/ penumonia_label = 1 = penumonia
            
#         if binary_label == 0:
#             pre_label = binary_label + 1   # correct_label = 1 = normal/ correct_label = 2 = penumonia
#         if binary_label == 1:
#             pre_label = binary_label + 1

#         covid_outputs = covid_model(inputs)
#         _, covid_preds = torch.max(covid_outputs, 1)
#         covid_confidence = torch.nn.functional.softmax(covid_outputs, dim=1)
#         covid_label = covid_preds.item()
#         covid_confidence_score = covid_confidence[0][0].item()

#         if covid_label == 0:
#             if covid_confidence_score > covid_threshold:
#                 pre_label = 0 
                
        # Basic OVR model
        covid_outputs = covid_model(inputs)
        _, covid_preds = torch.max(covid_outputs, 1)
        covid_label = covid_preds.item()
        covid_confidence = torch.nn.functional.softmax(covid_outputs, dim=1)
        covid_confidence_score = covid_confidence[0][0].item()
        
        pneumonia_outputs = pneumonia_model(inputs)
        _, pneumonia_preds = torch.max(pneumonia_outputs, 1)
        pneumonia_label = pneumonia_preds.item()
        pneumonia_confidence = torch.nn.functional.softmax(pneumonia_outputs, dim=1)
        pneumonia_confidence_score = pneumonia_confidence[0][0].item()
        
        normal_outputs = normal_model(inputs)
        _, normal_preds = torch.max(normal_outputs, 1)
        normal_label = normal_preds.item()
        normal_confidence = torch.nn.functional.softmax(normal_outputs, dim=1)
        normal_confidence_score = normal_confidence[0][0].item()
             
#         threshold = 0.95
        confidence_scores = [covid_confidence_score, normal_confidence_score, pneumonia_confidence_score]
        pre_label = np.argmax(confidence_scores)
        print(pre_label, confidence_scores)
                
        # Calculate metrics
        if pre_label == correct_label: 
            #Covid-index:0, Normal-index:1, Pneumonia-index:2
            running_corrects += 1 
            if pre_label == 0: 
                val_covid_TP += 1 #COVID 관점(Covid->Positive)에서는 covid를 정확히 분류하는 것이 TP
                val_normal_TN += 1 #Normal 관점에서는 covid를 정확히 분류하는 것이 TN
                val_pneumonia_TN += 1 #Pneumonia 관점에서는 covid를 정확히 분류하는 것이 TN

            elif pre_label == 1: 
                val_normal_TP += 1 #Normal 관점에서는 normal를 정확히 분류하는 것이 TP
                val_covid_TN += 1 #COVID 관점에서는 normal를 정확히 분류하는 것이 TN
                val_pneumonia_TN += 1 #Pneumonia 관점에서는 normal를 정확히 분류하는 것이 TN

            elif pre_label == 2: 
                val_pneumonia_TP += 1 #Pneumonia 관점에서는 pneumonia를 정확히 분류하는 것이 TP
                val_covid_TN += 1 #COVID 관점에서는 pneumonia를 정확히 분류하는 것이 TN
                val_normal_TN += 1 #COVID 관점에서는 pneumonia를 정확히 분류하는 것이 TN

        elif pre_label != correct_label:
            if pre_label == 0:
                if correct_label == 1:
                    val_covid_FP += 1 #COVID 관점(Covid->Positive)에서 covid라고 분류했지만 실제로는 normal인 경우는 FP
                    val_normal_FN += 1 #Normal 관점에서 covid라고 분류했지만 실제로는 normal인 경우는 FN
                    val_pneumonia_TN += 1 #<--추가 
                    #TN --> preds:0, label:1   
                    #--> Pneumonia 관점에서는 covid, normal 모두 Negative이다. 
                    #--> 즉, pred:0 --> covid --> Negative/ label:1 --> normal --> Negative
                    #--> 결과적으로 Penumonia관점에서는 실제 Negative(=normal)를 Negative(=covid)라고 분류했기 때문에 True가 된다.
                elif correct_label == 2:
                    val_covid_FP += 1 #COVID 관점에서 covid라고 분류했지만 실제로는 pneumonia인 경우는 FP
                    val_pneumonia_FN += 1 #Pneumonia 관점에서 covid라고 분류했지만 실제로는 pneumonia인 경우는 FN
                    val_normal_TN += 1 #<--추가

            elif pre_label == 1:
                if correct_label == 0:
                    val_normal_FP += 1 #Normal 관점에서 normal이라고 분류했지만 실제로는 covid인 경우는 FP
                    val_covid_FN += 1 #COVID 관점에서 normal이라고 분류했지만 실제로는 covid인 경우는 FN
                    val_pneumonia_TN += 1 #<--추가
                elif correct_label == 2:
                    val_normal_FP += 1 #NORMAL 관점에서 normal이라고 분류했지만 실제로는 pneumonia인 경우는 FP
                    val_pneumonia_FN += 1 #Pneumonia 관점에서 normal라고 분류했지만 실제로는 pneumonia인 경우는 FN
                    val_covid_TN += 1 #<-- 추가

            elif pre_label == 2:
                if correct_label == 0:
                    val_pneumonia_FP += 1 #Pneumonia 관점에서 pneumonia라고 분류했지만 실제로는 covid인 경우는 FP
                    val_covid_FN += 1 #COVID 관점에서 pneumonia라고 분류했지만 실제로는 covid인 경우는 FN
                    val_normal_TN += 1 #<-- 추가
                elif correct_label == 1:
                    val_pneumonia_FP += 1 #Pneumonia 관점에서 pneumonia라고 분류했지만 실제로는 normal인 경우는 FP
                    val_normal_FN += 1 #Normal 관점에서 pneumonia라고 분류했지만 실제로는 normal인 경우는 FN
                    val_covid_TN += 1 #<--추가

    #recall, precision -> https://en.wikipedia.org/wiki/Precision_and_recall
    #recall = TP/(TP+FN)
    #precision = TP/(TP+FP)
    '''
    if val_false_TP + val_false_FP == 0:
        val_false_FP = 0.000001
    '''    
        
    #Recall
    covid_recall = round(val_covid_TP/(val_covid_TP + val_covid_FN),4)*100
    normal_recall = round(val_normal_TP/(val_normal_TP + val_normal_FN),4)*100
    pneumonia_recall = round(val_pneumonia_TP/(val_pneumonia_TP + val_pneumonia_FN),4)*100

    #Preicision
    covid_precision = round(val_covid_TP/(val_covid_TP + val_covid_FP),4)*100
    normal_precision = round(val_normal_TP/(val_normal_TP + val_normal_FP),4)*100
    pneumonia_precision = round(val_pneumonia_TP/(val_pneumonia_TP + val_pneumonia_FP),4)*100
        
    #ACC: Accuracy
    covid_acc = round((val_covid_TP + val_covid_TN)/(val_covid_TP + val_covid_FP +val_covid_TN + val_covid_FN),4)*100
    normal_acc = round((val_normal_TP + val_normal_TN)/(val_normal_TP + val_normal_FP +val_normal_TN + val_normal_FN),4)*100
    pneumonia_acc = round((val_pneumonia_TP + val_pneumonia_TN)/(val_pneumonia_TP + val_pneumonia_FP +val_pneumonia_TN + val_pneumonia_FN),4)*100
        

    '''
    if false_precision+false_recall == 0:
        false_precision = 0.000001
    '''
        
    #F1 score -> https://en.wikipedia.org/wiki/F-score
    #F1 score = 2/((1/recall)+(1/precision)) = 2((precision*recall)/(precision+recall)) = tp/(tp+((1/2)(fp+fn)))
    covid_f1_score = round(2*(covid_precision*covid_recall)/(covid_precision+covid_recall),2)
    normal_f1_score = round(2*(normal_precision*normal_recall)/(normal_precision+normal_recall),2)
    pneumonia_f1_score = round(2*(pneumonia_precision*pneumonia_recall)/(pneumonia_precision+pneumonia_recall),2)

    #total_data = batch size --> covid_total_dataset = false_total_dataset
    covid_total_dataset = int(val_covid_TP + val_covid_FP +val_covid_TN + val_covid_FN)
    normal_total_dataset = int(val_normal_TP + val_normal_FP +val_normal_TN + val_normal_FN)
    pneumonia_total_dataset = int(val_pneumonia_TP + val_pneumonia_FP +val_pneumonia_TN + val_pneumonia_FN)
        
    total_acc = running_corrects / covid_total_dataset
    
    # Write results to the work sheet
    ws['B2'] = total_acc
    
    ws['D2'] = covid_acc
    ws['E2'] = covid_recall
    ws['F2'] = covid_precision
    ws['G2'] = covid_f1_score
        
    ws['I2'] = normal_acc
    ws['J2'] = normal_recall
    ws['K2'] = normal_precision
    ws['L2'] = normal_f1_score

    ws['N2'] = pneumonia_acc
    ws['O2'] = pneumonia_recall
    ws['P2'] = pneumonia_precision
    ws['Q2'] = pneumonia_f1_score

      
    print('Val Acc: {:.4f}'.format(total_acc))    

    print()
    print('데이터셋A: ' + str(covid_total_dataset))
    print('Covid Recall: ' + str(covid_recall))        
    print('Covid Precision: ' + str(covid_precision))
    print('Covid ACC: ' + str(covid_acc))
    print('Covid F1 score: ' + str(covid_f1_score))
    print()
        
    print('데이터셋B: ' + str(normal_total_dataset))
    print('Normal Recall: ' + str(normal_recall))
    print('Normal Precision: ' + str(normal_precision))
    print('Normal ACC: ' + str(normal_acc))
    print('Normal F1 score: ' + str(normal_f1_score))
    print()
        
    print('데이터셋C: ' + str(pneumonia_total_dataset))
    print('Pneumonia Recall: ' + str(pneumonia_recall))
    print('Pneumonia Precision: ' + str(pneumonia_precision))
    print('Pneumonia ACC: ' + str(pneumonia_acc))
    print('Pneumonia F1 score: ' + str(pneumonia_f1_score))
        
    time_elapsed = time.time() - since
    print('Training complete in {:.0f}m {:.0f}s'.format(
        time_elapsed // 60, time_elapsed % 60))
    #print('Best val Acc: {:4f}'.format(best_acc))

    # load best model weights
    #model.load_state_dict(best_covid_wts)
    if not os.path.isdir(out_dir):
        os.mkdir(out_dir)
        
    wb.save(f'{out_dir}/integrate_test.xlsx') # 엑셀로 저장한다. 
    #torch.save(model.state_dict(), 'covid_binary.pt')